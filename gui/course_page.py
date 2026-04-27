import csv
from datetime import datetime
from PyQt5.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QHBoxLayout,
    QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QMessageBox, QHeaderView, QComboBox, QFileDialog, QDialog
)

from CORE.db import get_conn


class CoursePage(QWidget):

    def __init__(self):
        super().__init__()
        self.setObjectName("CoursePage")

        # =========================
        # 新建课程
        # =========================
        self.course_name_input = QLineEdit()
        self.course_name_input.setPlaceholderText("课程名称")

        self.start_time_input = QLineEdit()
        self.start_time_input.setPlaceholderText("开始时间 例如 2026-04-07 08:00:00")

        self.end_time_input = QLineEdit()
        self.end_time_input.setPlaceholderText("结束时间 例如 2026-04-07 09:40:00")

        self.btn_create = QPushButton("创建课程")
        self.btn_update = QPushButton("保存修改")
        self.btn_delete = QPushButton("删除课程")

        create_layout = QHBoxLayout()
        create_layout.addWidget(self.course_name_input)
        create_layout.addWidget(self.start_time_input)
        create_layout.addWidget(self.end_time_input)
        create_layout.addWidget(self.btn_create)
        create_layout.addWidget(self.btn_update)
        create_layout.addWidget(self.btn_delete)

        # =========================
        # 导入名单
        # =========================
        self.course_select = QComboBox()
        self.btn_import = QPushButton("导入课程名单 (CSV)")
        self.btn_refresh = QPushButton("刷新课程")
        self.btn_export = QPushButton("导出签到记录 (CSV)")
        self.btn_view_roster = QPushButton("查看课程成员")
        self.btn_remove_roster = QPushButton("删除课程成员")
        self.btn_add_from_users = QPushButton("从系统成员添加")

        import_layout = QHBoxLayout()
        import_layout.addWidget(QLabel("选择课程"))
        import_layout.addWidget(self.course_select)
        import_layout.addWidget(self.btn_import)
        import_layout.addWidget(self.btn_refresh)
        import_layout.addWidget(self.btn_export)
        import_layout.addWidget(self.btn_view_roster)
        import_layout.addWidget(self.btn_add_from_users)
        import_layout.addWidget(self.btn_remove_roster)

        # =========================
        # 课程列表
        # =========================
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["ID", "课程名称", "开始时间", "结束时间"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)

        layout = QVBoxLayout()
        layout.addLayout(create_layout)
        layout.addLayout(import_layout)
        layout.addWidget(self.table)
        self.setLayout(layout)

        # =========================
        # 信号
        # =========================
        self.btn_create.clicked.connect(self.create_course)
        self.btn_update.clicked.connect(self.update_course)
        self.btn_delete.clicked.connect(self.delete_course)
        self.btn_import.clicked.connect(self.import_roster)
        self.btn_refresh.clicked.connect(self.load_courses)
        self.btn_export.clicked.connect(self.export_attendance)
        self.btn_view_roster.clicked.connect(self.view_roster)
        self.btn_add_from_users.clicked.connect(self.add_from_users)
        self.btn_remove_roster.clicked.connect(self.remove_from_roster)

        self.btn_create.setStyleSheet("color: #2e7d32;")
        self.btn_update.setStyleSheet("color: #1976d2;")
        self.btn_delete.setStyleSheet("color: #d32f2f;")
        self.btn_import.setStyleSheet("color: #2e7d32;")
        self.btn_add_from_users.setStyleSheet("color: #2e7d32;")
        self.btn_remove_roster.setStyleSheet("color: #d32f2f;")
        self.setStyleSheet("""
        QWidget#CoursePage QPushButton {
            color: #1f2937;
            background-color: rgba(255, 255, 255, 240);
            border: 1px solid rgba(0, 0, 0, 50);
            border-radius: 6px;
        }
        QWidget#CoursePage QPushButton:hover {
            background-color: rgba(245, 245, 245, 245);
        }
        QWidget#CoursePage QPushButton:pressed {
            background-color: rgba(232, 232, 232, 245);
        }
        """)

        self.load_courses()
        self._selected_course_id = None
        self.table.itemSelectionChanged.connect(self._fill_form_from_selection)

    def load_courses(self):
        conn = get_conn()
        cur = conn.cursor()

        cur.execute("""
            SELECT id, name, start_time, end_time
            FROM courses
            ORDER BY start_time DESC
        """)
        rows = cur.fetchall()
        conn.close()

        self.table.setRowCount(len(rows))
        self.course_select.clear()

        for row_idx, (course_id, name, start_time, end_time) in enumerate(rows):
            self.table.setItem(row_idx, 0, QTableWidgetItem(str(course_id)))
            self.table.setItem(row_idx, 1, QTableWidgetItem(name))
            self.table.setItem(row_idx, 2, QTableWidgetItem(str(start_time)))
            self.table.setItem(row_idx, 3, QTableWidgetItem(str(end_time)))

            self.course_select.addItem(f"{course_id} - {name}", course_id)
        self._selected_course_id = None

    def create_course(self):
        name = self.course_name_input.text().strip()
        start_time = self.start_time_input.text().strip()
        end_time = self.end_time_input.text().strip()

        if name == "" or start_time == "" or end_time == "":
            QMessageBox.warning(self, "提示", "请完整填写课程名称、开始时间、结束时间")
            return

        try:
            start_dt = datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S")
            end_dt = datetime.strptime(end_time, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            QMessageBox.warning(self, "提示", "时间格式错误，请使用：YYYY-MM-DD HH:MM:SS")
            return

        if end_dt <= start_dt:
            QMessageBox.warning(self, "提示", "结束时间必须晚于开始时间")
            return

        conn = get_conn()
        cur = conn.cursor()

        cur.execute(
            "INSERT INTO courses (name, start_time, end_time) VALUES (%s,%s,%s)",
            (name, start_dt, end_dt)
        )

        conn.commit()
        conn.close()

        self.course_name_input.setText("")
        self.start_time_input.setText("")
        self.end_time_input.setText("")

        QMessageBox.information(self, "提示", "课程已创建")
        self.load_courses()

    def update_course(self):
        if self._selected_course_id is None:
            QMessageBox.warning(self, "提示", "请先选择要编辑的课程")
            return

        name = self.course_name_input.text().strip()
        start_time = self.start_time_input.text().strip()
        end_time = self.end_time_input.text().strip()

        if name == "" or start_time == "" or end_time == "":
            QMessageBox.warning(self, "提示", "请完整填写课程名称、开始时间、结束时间")
            return

        try:
            start_dt = datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S")
            end_dt = datetime.strptime(end_time, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            QMessageBox.warning(self, "提示", "时间格式错误，请使用：YYYY-MM-DD HH:MM:SS")
            return

        if end_dt <= start_dt:
            QMessageBox.warning(self, "提示", "结束时间必须晚于开始时间")
            return

        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "UPDATE courses SET name=%s, start_time=%s, end_time=%s WHERE id=%s",
            (name, start_dt, end_dt, self._selected_course_id)
        )
        conn.commit()
        conn.close()

        QMessageBox.information(self, "提示", "课程已更新")
        self.load_courses()

    def delete_course(self):
        if self._selected_course_id is None:
            QMessageBox.warning(self, "提示", "请先选择要删除的课程")
            return

        confirm = QMessageBox.question(
            self,
            "确认删除",
            "确认删除该课程吗？这会同时删除课程名单和签到记录。",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm != QMessageBox.Yes:
            return

        conn = get_conn()
        cur = conn.cursor()
        cur.execute("DELETE FROM course_roster WHERE course_id=%s", (self._selected_course_id,))
        cur.execute("DELETE FROM attendance WHERE course_id=%s", (self._selected_course_id,))
        cur.execute("DELETE FROM courses WHERE id=%s", (self._selected_course_id,))
        conn.commit()
        conn.close()

        QMessageBox.information(self, "提示", "课程已删除")
        self.load_courses()

    def _fill_form_from_selection(self):
        selected = self.table.selectionModel().selectedRows()
        if not selected:
            self._selected_course_id = None
            return

        row = selected[0].row()
        self._selected_course_id = int(self.table.item(row, 0).text())
        self.course_name_input.setText(self.table.item(row, 1).text())
        self.start_time_input.setText(self.table.item(row, 2).text())
        self.end_time_input.setText(self.table.item(row, 3).text())

    def import_roster(self):
        course_id = self.course_select.currentData()
        if course_id is None:
            QMessageBox.warning(self, "提示", "请先选择课程")
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择名单文件",
            "",
            "CSV Files (*.csv);;Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        student_ids = []
        if file_path.lower().endswith(".csv"):
            with open(file_path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                if "student_id" not in reader.fieldnames:
                    QMessageBox.warning(self, "提示", "CSV 需包含 student_id 列")
                    return
                student_ids = [row["student_id"].strip() for row in reader if row.get("student_id")]
        elif file_path.lower().endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except Exception:
                QMessageBox.warning(self, "提示", "缺少 openpyxl 依赖，请先安装：pip install openpyxl")
                return

            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
            if "student_id" not in header:
                QMessageBox.warning(self, "提示", "Excel 需包含 student_id 列")
                return

            sid_idx = header.index("student_id") + 1

            for row in ws.iter_rows(min_row=2):
                cell = row[sid_idx - 1]
                if cell.value is None:
                    continue
                student_ids.append(str(cell.value).strip())
        else:
            QMessageBox.warning(self, "提示", "不支持的文件类型")
            return

        if not student_ids:
            QMessageBox.warning(self, "提示", "CSV 未读取到 student_id")
            return

        conn = get_conn()
        cur = conn.cursor()

        # 查找用户ID
        format_strings = ",".join(["%s"] * len(student_ids))
        cur.execute(
            f"SELECT id, student_id FROM users WHERE student_id IN ({format_strings})",
            tuple(student_ids)
        )
        found = cur.fetchall()

        user_id_map = {sid: uid for uid, sid in found}

        inserted = 0
        missing = 0

        for sid in student_ids:
            user_id = user_id_map.get(sid)
            if not user_id:
                missing += 1
                continue

            cur.execute(
                "SELECT 1 FROM course_roster WHERE course_id=%s AND user_id=%s",
                (course_id, user_id)
            )
            if cur.fetchone():
                continue

            cur.execute(
                "INSERT INTO course_roster (course_id, user_id) VALUES (%s,%s)",
                (course_id, user_id)
            )
            inserted += 1

        conn.commit()
        conn.close()

        QMessageBox.information(
            self,
            "导入完成",
            f"成功导入 {inserted} 人，未匹配到用户 {missing} 人"
        )

    def export_attendance(self):
        course_id = self.course_select.currentData()

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出签到记录",
            "attendance.csv",
            "CSV Files (*.csv);;Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        conn = get_conn()
        cur = conn.cursor()

        if course_id is None:
            cur.execute("""
                SELECT a.id, a.course_id, a.course_name, a.user_id,
                       u.name, u.student_id, u.class_name, u.major,
                       a.checkin_time
                FROM attendance a
                LEFT JOIN users u ON a.user_id = u.id
                ORDER BY a.checkin_time DESC
            """)
            rows = cur.fetchall()
        else:
            # 全部导出：课程名单 + 签到信息（含未签到）
            cur.execute("""
                SELECT r.user_id, u.name, u.student_id, u.class_name, u.major
                FROM course_roster r
                JOIN users u ON r.user_id = u.id
                WHERE r.course_id = %s
                ORDER BY u.student_id
            """, (course_id,))
            roster = cur.fetchall()

            cur.execute("""
                SELECT a.user_id, a.course_name, a.checkin_time
                FROM attendance a
                WHERE a.course_id = %s
            """, (course_id,))
            present_rows = cur.fetchall()
            present_map = {uid: (course_name, checkin_time) for uid, course_name, checkin_time in present_rows}

            rows = []
            for user_id, name, student_id, class_name, major in roster:
                if user_id in present_map:
                    course_name, checkin_time = present_map[user_id]
                    rows.append((course_id, course_name, user_id, name, student_id, class_name, major, checkin_time, "已签到"))
                else:
                    rows.append((course_id, "", user_id, name, student_id, class_name, major, "", "未签到"))
        conn.close()

        if course_id is None:
            headers = [
                "attendance_id",
                "course_id",
                "course_name",
                "user_id",
                "name",
                "student_id",
                "class_name",
                "major",
                "checkin_time"
            ]
        else:
            headers = [
                "course_id",
                "course_name",
                "user_id",
                "name",
                "student_id",
                "class_name",
                "major",
                "checkin_time",
                "status"
            ]

        if file_path.lower().endswith(".xlsx"):
            try:
                from openpyxl import Workbook
            except Exception:
                QMessageBox.warning(self, "提示", "缺少 openpyxl 依赖，请先安装：pip install openpyxl")
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "attendance"
            ws.append(headers)
            for row in rows:
                ws.append(list(row))
            wb.save(file_path)
        else:
            if not file_path.lower().endswith(".csv"):
                file_path += ".csv"
            with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for row in rows:
                    writer.writerow(row)

        QMessageBox.information(self, "提示", f"已导出 {len(rows)} 条签到记录")

    def view_roster(self):
        course_id = self.course_select.currentData()
        if course_id is None:
            QMessageBox.warning(self, "提示", "请先选择课程")
            return

        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT u.id, u.name, u.student_id, u.class_name, u.major
            FROM course_roster r
            JOIN users u ON r.user_id = u.id
            WHERE r.course_id = %s
            ORDER BY u.student_id
        """, (course_id,))
        rows = cur.fetchall()
        conn.close()

        dialog = QDialog(self)
        dialog.setWindowTitle("课程成员")
        layout = QVBoxLayout()

        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["ID", "姓名", "学号", "班级", "专业"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setRowCount(len(rows))

        for row_idx, (user_id, name, student_id, class_name, major) in enumerate(rows):
            table.setItem(row_idx, 0, QTableWidgetItem(str(user_id)))
            table.setItem(row_idx, 1, QTableWidgetItem(name or ""))
            table.setItem(row_idx, 2, QTableWidgetItem(student_id or ""))
            table.setItem(row_idx, 3, QTableWidgetItem(class_name or ""))
            table.setItem(row_idx, 4, QTableWidgetItem(major or ""))

        layout.addWidget(table)
        dialog.setLayout(layout)
        dialog.resize(700, 400)
        dialog.exec_()

    def add_from_users(self):
        course_id = self.course_select.currentData()
        if course_id is None:
            QMessageBox.warning(self, "提示", "请先选择课程")
            return

        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, name, student_id, class_name, major
            FROM users
            ORDER BY student_id
        """)
        rows = cur.fetchall()
        conn.close()

        dialog = QDialog(self)
        dialog.setWindowTitle("选择成员加入课程")
        layout = QVBoxLayout()

        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["ID", "姓名", "学号", "班级", "专业"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setRowCount(len(rows))
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setSelectionMode(QTableWidget.MultiSelection)

        for row_idx, (user_id, name, student_id, class_name, major) in enumerate(rows):
            table.setItem(row_idx, 0, QTableWidgetItem(str(user_id)))
            table.setItem(row_idx, 1, QTableWidgetItem(name or ""))
            table.setItem(row_idx, 2, QTableWidgetItem(student_id or ""))
            table.setItem(row_idx, 3, QTableWidgetItem(class_name or ""))
            table.setItem(row_idx, 4, QTableWidgetItem(major or ""))

        btn_add = QPushButton("添加所选成员")
        btn_cancel = QPushButton("取消")

        btn_layout = QHBoxLayout()
        btn_layout.addWidget(btn_add)
        btn_layout.addWidget(btn_cancel)

        layout.addWidget(table)
        layout.addLayout(btn_layout)
        dialog.setLayout(layout)
        dialog.resize(700, 450)

        def do_add():
            selected = table.selectionModel().selectedRows()
            if not selected:
                QMessageBox.warning(self, "提示", "请先选择成员")
                return

            user_ids = [table.item(idx.row(), 0).text() for idx in selected]

            conn2 = get_conn()
            cur2 = conn2.cursor()

            inserted = 0
            for uid in user_ids:
                cur2.execute(
                    "SELECT 1 FROM course_roster WHERE course_id=%s AND user_id=%s",
                    (course_id, uid)
                )
                if cur2.fetchone():
                    continue
                cur2.execute(
                    "INSERT INTO course_roster (course_id, user_id) VALUES (%s,%s)",
                    (course_id, uid)
                )
                inserted += 1

            conn2.commit()
            conn2.close()

            QMessageBox.information(self, "提示", f"已添加 {inserted} 人")
            dialog.accept()

        btn_add.clicked.connect(do_add)
        btn_cancel.clicked.connect(dialog.reject)

        dialog.exec_()

    def remove_from_roster(self):
        course_id = self.course_select.currentData()
        if course_id is None:
            QMessageBox.warning(self, "提示", "请先选择课程")
            return

        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT u.id, u.name, u.student_id, u.class_name, u.major
            FROM course_roster r
            JOIN users u ON r.user_id = u.id
            WHERE r.course_id = %s
            ORDER BY u.student_id
        """, (course_id,))
        rows = cur.fetchall()
        conn.close()

        dialog = QDialog(self)
        dialog.setWindowTitle("删除课程成员")
        layout = QVBoxLayout()

        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["ID", "姓名", "学号", "班级", "专业"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setRowCount(len(rows))
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setSelectionMode(QTableWidget.MultiSelection)

        for row_idx, (user_id, name, student_id, class_name, major) in enumerate(rows):
            table.setItem(row_idx, 0, QTableWidgetItem(str(user_id)))
            table.setItem(row_idx, 1, QTableWidgetItem(name or ""))
            table.setItem(row_idx, 2, QTableWidgetItem(student_id or ""))
            table.setItem(row_idx, 3, QTableWidgetItem(class_name or ""))
            table.setItem(row_idx, 4, QTableWidgetItem(major or ""))

        btn_remove = QPushButton("删除所选成员")
        btn_cancel = QPushButton("取消")

        btn_layout = QHBoxLayout()
        btn_layout.addWidget(btn_remove)
        btn_layout.addWidget(btn_cancel)

        layout.addWidget(table)
        layout.addLayout(btn_layout)
        dialog.setLayout(layout)
        dialog.resize(700, 450)

        def do_remove():
            selected = table.selectionModel().selectedRows()
            if not selected:
                QMessageBox.warning(self, "提示", "请先选择成员")
                return

            confirm = QMessageBox.question(
                self,
                "确认删除",
                "确认从课程名单中删除所选成员吗？",
                QMessageBox.Yes | QMessageBox.No
            )
            if confirm != QMessageBox.Yes:
                return

            user_ids = [table.item(idx.row(), 0).text() for idx in selected]

            conn2 = get_conn()
            cur2 = conn2.cursor()
            deleted = 0

            for uid in user_ids:
                cur2.execute(
                    "DELETE FROM course_roster WHERE course_id=%s AND user_id=%s",
                    (course_id, uid)
                )
                deleted += cur2.rowcount

            conn2.commit()
            conn2.close()

            QMessageBox.information(self, "提示", f"已删除 {deleted} 人")
            dialog.accept()

        btn_remove.clicked.connect(do_remove)
        btn_cancel.clicked.connect(dialog.reject)

        dialog.exec_()
